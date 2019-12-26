import os
import io
import openpyxl
import csv
from Account import Account
from Project import Project


class Excel_Reader:

    def __init__(self, month, year, newPath):
        self.month = month
        self.year = year
        self.path = newPath
        self.excel_files = self.get_excel_files(newPath+'\\'+str(year)+'\\'+str(month))



    def get_path(self):
        fileDir = os.path.dirname(os.path.abspath(__file__))
        print(fileDir)
        parentDir = os.path.dirname(fileDir)
        newPath = os.path.join(parentDir, 'IS')
        print(newPath)
        return newPath


    def get_excel_files(self, path):
        files = []

        for r, d, f in os.walk(path):
            for file in f:
                if '.xlsx' in file:
                    files.append(os.path.join(r, file))


        return files


    def reading_excels(self):
        path = self.path+'\\'+str(self.month)+str(self.year)+'result.csv'
        print(path)
        with open(path, 'w', newline='') as csvfile:
            fieldnames = ['User Id','Month Date', 'Vendor', 'Project Name', 'Hours Poland',
                    'Hours Abroad', 'Total hours', 'Man days', 'Sick Leaves', 'Other absence']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()

            for f in self.excel_files:
                wb = openpyxl.load_workbook(f, data_only=True)

                sheet = wb['202001']


                user_id = sheet['A2'].value
                vendor = sheet['A4'].value
                date = sheet['B4'].value
                total = sheet['E35'].value
                man_days = sheet['E36'].value
                work_from_poland = sheet['E37'].value
                work_from_abroad = sheet['E38'].value
                sick_leaves = sheet['E39'].value
                other = sheet['E40'].value

                account = Account(user_id, vendor, date, sick_leaves,
                                    other, total, man_days, work_from_poland,
                                     work_from_abroad)

                for i in range(6,10):
                    name_of_project = sheet.cell(row=3, column=i).value

                    if name_of_project == 'N/A' or name_of_project == '':
                        break

                    total_project = sheet.cell(row=35, column=i).value
                    man_days_project = sheet.cell(row=36, column=i).value
                    work_from_poland_project = sheet.cell(row=37, column=i).value
                    work_from_abroad_project = sheet.cell(row=38, column=i).value


                    account.add_project(Project(name_of_project,
                                                work_from_abroad_project,
                                                work_from_poland_project,
                                                total_project,
                                                man_days_project
                                                ))

                for project in account.get_projects():

                    writer.writerow({'User Id': int(account.get_id_account()),
                                    'Month Date': account.get_date().date(),
                                    'Vendor': account.get_vendor(),
                                    'Project Name': project.get_name_of_project(),
                                    'Hours Poland': project.get_work_from_poland(),
                                    'Hours Abroad':project.get_work_from_abroad(),
                                    'Total hours':project.get_total(),
                                    'Man days':project.get_man_days(),
                                    'Sick Leaves':account.get_sick_leaves(),
                                    'Other absence': account.get_other()})
